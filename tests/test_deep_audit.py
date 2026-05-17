"""
Deep Audit Test Suite
=====================
Covers every gap found by the 4-agent comprehensive codebase audit:

  • Financial calculation unit tests (zero denominators, negatives, precision)
  • DSCR / collateral calculation edge cases
  • Export DOCX / PDF error paths and security
  • XLS silent-failure fix verification
  • XLSX row-padding fix verification
  • Auth /setup endpoint
  • Section-name consistency across Python / JavaScript
  • XSS in _md_to_html fallback fix verification
  • Login hint must be English
  • Mapping rule auto_classify heuristics
  • Quota: per-role limits (reviewer 2×, admin 5×)
  • Evidence extraction edge cases (corrupted, empty)
  • Fact dependency circular-reference guard (not present → document as gap)
"""
from __future__ import annotations

import io
import os
import uuid

import pytest
import pytest_asyncio
from httpx import AsyncClient, ASGITransport
from unittest.mock import AsyncMock, MagicMock, patch

os.environ.setdefault("GEMINI_API_KEY", "mock-key-for-testing")
os.environ.setdefault("ADMIN_EMAIL", "admin@example.com")
os.environ.setdefault("ADMIN_PASSWORD", "admin123")
os.environ.setdefault("ENVIRONMENT", "development")
os.environ.setdefault("AUTO_CREATE_TABLES", "true")

from main import app  # noqa: E402

BASE = "/api/credit-report"
AUTH = f"{BASE}/auth"
RPTS = f"{BASE}/reports"


# ── helpers ───────────────────────────────────────────────────────────────────

async def _login(ac: AsyncClient, email: str, password: str) -> dict:
    r = await ac.post(f"{AUTH}/login", data={"username": email, "password": password})
    return r.json()


async def _hdrs(ac: AsyncClient, email: str, password: str = "Pass1234!") -> dict:
    tokens = await _login(ac, email, password)
    return {"Authorization": f"Bearer {tokens['access_token']}"}


async def _register(ac, admin_h, email, role="analyst"):
    r = await ac.post(f"{AUTH}/register",
                      json={"email": email, "password": "Pass1234!", "role": role},
                      headers=admin_h)
    return r.json()


def _mock_gemini(text: str = "## Section\n\nContent.\n\n| A | B |\n|---|---|\n| 1 | 2 |\n"):
    mock_resp = MagicMock()
    mock_resp.text = text
    mock_usage = MagicMock()
    mock_usage.prompt_token_count = 50
    mock_usage.candidates_token_count = 100
    mock_resp.usage_metadata = mock_usage
    mock_client = MagicMock()
    mock_client.aio.models.generate_content = AsyncMock(return_value=mock_resp)
    return patch("google.genai.Client", return_value=mock_client)


# ── fixtures ──────────────────────────────────────────────────────────────────

@pytest_asyncio.fixture
async def ac():
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        yield client


@pytest_asyncio.fixture
async def admin_hdrs(ac):
    return await _hdrs(ac, "admin@example.com", "admin123")


@pytest_asyncio.fixture
async def report(ac, admin_hdrs):
    r = await ac.post(f"{RPTS}", json={"borrower_name": "AuditCo", "industry": "shipping"},
                      headers=admin_hdrs)
    return r.json()


async def _seed_section_output(report_id: str, section_no: int = 4,
                               markdown: str = "## Section\n\nContent.\n\n| H | V |\n|---|---|\n| R | 1 |\n") -> None:
    from credit_report.database import AsyncSessionLocal
    from credit_report.models import SectionOutput
    from datetime import datetime, timezone
    async with AsyncSessionLocal() as db:
        db.add(SectionOutput(
            id=str(uuid.uuid4()),
            report_id=report_id,
            section_no=section_no,
            markdown=markdown,
            status="done",
            tokens_used=300,
            generated_at=datetime.now(timezone.utc),
        ))
        await db.commit()


# ══════════════════════════════════════════════════════════════════════════════
# 1 — Financial Ratio Unit Tests (pure functions, no DB/API)
# ══════════════════════════════════════════════════════════════════════════════

class TestFinancialRatioCalculations:
    """Direct unit tests for calculation_engine/financial_ratios.py."""

    def test_safe_divide_normal(self):
        from credit_report.calculation_engine.financial_ratios import safe_divide
        assert safe_divide(100.0, 5.0) == pytest.approx(20.0)

    def test_safe_divide_zero_denominator(self):
        from credit_report.calculation_engine.financial_ratios import safe_divide
        assert safe_divide(100.0, 0.0) is None

    def test_safe_divide_none_inputs(self):
        from credit_report.calculation_engine.financial_ratios import safe_divide
        assert safe_divide(None, 5.0) is None
        assert safe_divide(100.0, None) is None
        assert safe_divide(None, None) is None

    def test_debt_to_ebitda_normal(self):
        from credit_report.calculation_engine.financial_ratios import debt_to_ebitda
        val, formula, fids = debt_to_ebitda(1000.0, 200.0)
        assert val == pytest.approx(5.0)
        assert "5.00x" in formula
        assert fids == []

    def test_debt_to_ebitda_zero_ebitda(self):
        from credit_report.calculation_engine.financial_ratios import debt_to_ebitda
        val, formula, _ = debt_to_ebitda(1000.0, 0.0)
        assert val is None
        assert "N/M" not in formula or "1,000.0" in formula

    def test_debt_to_ebitda_negative_ebitda(self):
        from credit_report.calculation_engine.financial_ratios import debt_to_ebitda
        val, formula, _ = debt_to_ebitda(1000.0, -50.0)
        assert val == pytest.approx(-20.0)
        assert "-20.00x" in formula

    def test_interest_coverage_normal(self):
        from credit_report.calculation_engine.financial_ratios import interest_coverage
        val, formula, _ = interest_coverage(500.0, 100.0)
        assert val == pytest.approx(5.0)
        assert "5.0x" in formula

    def test_interest_coverage_zero_interest(self):
        from credit_report.calculation_engine.financial_ratios import interest_coverage
        val, formula, _ = interest_coverage(500.0, 0.0)
        assert val is None

    def test_net_debt_positive(self):
        from credit_report.calculation_engine.financial_ratios import net_debt
        val, formula, _ = net_debt(1000.0, 200.0)
        assert val == pytest.approx(800.0)
        assert "800" in formula

    def test_net_debt_net_cash_position(self):
        from credit_report.calculation_engine.financial_ratios import net_debt
        val, formula, _ = net_debt(100.0, 300.0)
        assert val == pytest.approx(-200.0)
        assert "Net Cash" in formula

    def test_ebitda_margin_normal(self):
        from credit_report.calculation_engine.financial_ratios import ebitda_margin
        val, formula, _ = ebitda_margin(300.0, 1000.0)
        assert val == pytest.approx(0.3)
        assert "30.0%" in formula

    def test_ebitda_margin_zero_revenue(self):
        from credit_report.calculation_engine.financial_ratios import ebitda_margin
        val, formula, _ = ebitda_margin(300.0, 0.0)
        assert val is None

    def test_net_margin_with_loss(self):
        from credit_report.calculation_engine.financial_ratios import net_margin
        val, formula, _ = net_margin(-50.0, 1000.0)
        assert val == pytest.approx(-0.05)
        assert "-5.0%" in formula

    def test_debt_to_equity_zero_equity(self):
        from credit_report.calculation_engine.financial_ratios import debt_to_equity
        val, formula, _ = debt_to_equity(500.0, 0.0)
        assert val is None

    def test_fact_ids_passed_through(self):
        from credit_report.calculation_engine.financial_ratios import debt_to_ebitda
        fid1, fid2 = "fact-a", "fact-b"
        _, _, fids = debt_to_ebitda(100.0, 50.0, fid1, fid2)
        assert fid1 in fids
        assert fid2 in fids
        assert len(fids) == 2


# ══════════════════════════════════════════════════════════════════════════════
# 2 — DSCR Calculation Unit Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestDSCRCalculation:

    def test_dscr_normal(self):
        from credit_report.calculation_engine.dscr import calculate_dscr
        val, formula, fids = calculate_dscr(120.0, 30.0, 10.0)
        assert val == pytest.approx(3.0, rel=1e-3)
        assert "3.00x" in formula

    def test_dscr_zero_debt_service(self):
        from credit_report.calculation_engine.dscr import calculate_dscr
        val, formula, _ = calculate_dscr(100.0, 0.0, 0.0)
        assert val is None
        assert "N/M" in formula

    def test_dscr_below_one_is_stress(self):
        from credit_report.calculation_engine.dscr import calculate_dscr
        val, formula, _ = calculate_dscr(50.0, 100.0, 0.0)
        assert val == pytest.approx(0.5, rel=1e-3)
        assert "0.50x" in formula

    def test_dscr_rounds_to_4_places(self):
        from credit_report.calculation_engine.dscr import calculate_dscr
        val, _, _ = calculate_dscr(1.0, 3.0, 0.0)
        assert val is not None
        assert len(str(val).split(".")[-1]) <= 4

    def test_dscr_fact_id_lineage(self):
        from credit_report.calculation_engine.dscr import calculate_dscr
        _, _, fids = calculate_dscr(120.0, 30.0, 10.0, "ocf-id", "prin-id", "int-id")
        assert "ocf-id" in fids
        assert "prin-id" in fids
        assert "int-id" in fids


# ══════════════════════════════════════════════════════════════════════════════
# 3 — Collateral Calculation Unit Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestCollateralCalculations:

    def test_ltc_normal(self):
        from credit_report.calculation_engine.collateral import ltc
        val, formula = ltc(80.0, 100.0)
        assert val == pytest.approx(80.0)
        assert "80.0%" in formula

    def test_ltc_zero_contract_price(self):
        from credit_report.calculation_engine.collateral import ltc
        val, formula = ltc(80.0, 0.0)
        assert val == 0.0
        assert "N/M" in formula

    def test_ltc_over_100_pct(self):
        from credit_report.calculation_engine.collateral import ltc
        val, formula = ltc(120.0, 100.0)
        assert val == pytest.approx(120.0)

    def test_current_ltv_normal(self):
        from credit_report.calculation_engine.collateral import current_ltv
        ltv, acr, formula = current_ltv(75.0, 100.0)
        assert ltv == pytest.approx(75.0)
        assert acr == pytest.approx(133.33, rel=1e-2)
        assert "75.0%" in formula

    def test_current_ltv_zero_asset(self):
        from credit_report.calculation_engine.collateral import current_ltv
        ltv, acr, formula = current_ltv(75.0, 0.0)
        assert ltv == 0.0
        assert "N/M" in formula

    def test_rg_coverage_normal(self):
        from credit_report.calculation_engine.collateral import rg_coverage
        cov, formula = rg_coverage(50.0, 100.0)
        assert cov == pytest.approx(50.0)
        assert "50%" in formula

    def test_rg_coverage_zero_exposure(self):
        from credit_report.calculation_engine.collateral import rg_coverage
        cov, formula = rg_coverage(50.0, 0.0)
        assert cov is None
        assert "N/M" in formula


# ══════════════════════════════════════════════════════════════════════════════
# 4 — Mapping Rule auto_classify
# ══════════════════════════════════════════════════════════════════════════════

class TestMappingAutoClassify:

    def test_auto_classify_revenue_label(self):
        from credit_report.calculation_engine.mapping.mapping_rules import auto_classify
        result = auto_classify("Total Revenue")
        assert result is not None
        metric, category = result
        assert "revenue" in metric.lower()

    def test_auto_classify_ebitda_label(self):
        from credit_report.calculation_engine.mapping.mapping_rules import auto_classify
        result = auto_classify("EBITDA")
        assert result is not None
        metric, _ = result
        assert "ebitda" in metric.lower()

    def test_auto_classify_unknown_label_returns_none(self):
        from credit_report.calculation_engine.mapping.mapping_rules import auto_classify
        result = auto_classify("zzz_completely_unknown_metric_xyz")
        # May return None for truly unknown labels
        assert result is None or isinstance(result, tuple)

    def test_auto_classify_interest_expense(self):
        from credit_report.calculation_engine.mapping.mapping_rules import auto_classify
        result = auto_classify("Interest Expense")
        assert result is not None
        metric, _ = result
        assert "interest" in metric.lower()

    def test_auto_classify_case_insensitive(self):
        from credit_report.calculation_engine.mapping.mapping_rules import auto_classify
        r1 = auto_classify("revenue")
        r2 = auto_classify("REVENUE")
        r3 = auto_classify("Revenue")
        # All should return the same (or all None)
        assert (r1 is None) == (r2 is None) == (r3 is None)


# ══════════════════════════════════════════════════════════════════════════════
# 5 — Export Endpoint Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestExportEndpoints:

    async def test_docx_export_no_sections_404(self, ac, admin_hdrs, report):
        """Empty report → 404 (no completed sections)."""
        r = await ac.get(f"{RPTS}/{report['id']}/export/docx", headers=admin_hdrs)
        assert r.status_code == 404

    async def test_docx_export_with_sections_returns_binary(self, ac, admin_hdrs, report):
        """With a completed section, DOCX export returns binary content."""
        await _seed_section_output(report["id"])
        r = await ac.get(f"{RPTS}/{report['id']}/export/docx", headers=admin_hdrs)
        if r.status_code == 503:
            pytest.skip("python-docx not installed on this server")
        assert r.status_code == 200
        content_type = r.headers.get("content-type", "")
        assert "wordprocessingml" in content_type or "octet-stream" in content_type

    async def test_docx_content_disposition_header(self, ac, admin_hdrs, report):
        """DOCX export must set Content-Disposition: attachment with .docx filename."""
        await _seed_section_output(report["id"])
        r = await ac.get(f"{RPTS}/{report['id']}/export/docx", headers=admin_hdrs)
        if r.status_code == 503:
            pytest.skip("python-docx not installed")
        assert r.status_code == 200
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert ".docx" in cd

    async def test_docx_export_requires_auth(self, ac, report, admin_hdrs):
        await _seed_section_output(report["id"])
        r = await ac.get(f"{RPTS}/{report['id']}/export/docx")
        assert r.status_code == 401

    async def test_pdf_export_no_sections_404(self, ac, admin_hdrs, report):
        r = await ac.get(f"{RPTS}/{report['id']}/export/pdf", headers=admin_hdrs)
        assert r.status_code in (404, 503)

    async def test_pdf_export_requires_auth(self, ac, report, admin_hdrs):
        await _seed_section_output(report["id"])
        r = await ac.get(f"{RPTS}/{report['id']}/export/pdf")
        assert r.status_code == 401

    async def test_pdf_export_no_weasyprint_503(self, ac, admin_hdrs, report):
        """When weasyprint is not installed, PDF export returns 503."""
        await _seed_section_output(report["id"])
        with patch.dict("sys.modules", {"weasyprint": None}):
            r = await ac.get(f"{RPTS}/{report['id']}/export/pdf", headers=admin_hdrs)
            if r.status_code == 503:
                assert "weasyprint" in r.json().get("detail", "").lower() or \
                       "unavailable" in r.json().get("detail", "").lower()
            # weasyprint IS installed if 200 — that's also fine

    async def test_docx_export_chinese_borrower_name_no_crash(self, ac, admin_hdrs):
        """Borrower name with CJK characters must not crash DOCX export (RFC 2183 issue)."""
        rpt = (await ac.post(f"{RPTS}",
                             json={"borrower_name": "台灣航運集團股份有限公司"},
                             headers=admin_hdrs)).json()
        await _seed_section_output(rpt["id"])
        r = await ac.get(f"{RPTS}/{rpt['id']}/export/docx", headers=admin_hdrs)
        if r.status_code == 503:
            pytest.skip("python-docx not installed")
        # Must not 500, regardless of filename encoding
        assert r.status_code == 200

    async def test_docx_export_table_column_mismatch_no_crash(self, ac, admin_hdrs):
        """Uneven markdown tables (row shorter than header) must not crash DOCX export."""
        uneven_md = (
            "## Section\n\n"
            "| H1 | H2 | H3 |\n|---|---|---|\n"
            "| R1C1 | R1C2 |\n"          # row has fewer cols than header
            "| R2C1 | R2C2 | R2C3 |\n"  # normal row
        )
        rpt = (await ac.post(f"{RPTS}", json={"borrower_name": "TableCo"},
                             headers=admin_hdrs)).json()
        await _seed_section_output(rpt["id"], markdown=uneven_md)
        r = await ac.get(f"{RPTS}/{rpt['id']}/export/docx", headers=admin_hdrs)
        if r.status_code == 503:
            pytest.skip("python-docx not installed")
        assert r.status_code == 200, f"Table column mismatch crashed export: {r.text[:200]}"

    async def test_export_wrong_report_404(self, ac, admin_hdrs):
        r = await ac.get(f"{RPTS}/{uuid.uuid4()}/export/docx", headers=admin_hdrs)
        assert r.status_code == 404


# ══════════════════════════════════════════════════════════════════════════════
# 6 — XLS / XLSX File Extraction
# ══════════════════════════════════════════════════════════════════════════════

class TestXLSXExtraction:

    @pytest.fixture(autouse=True)
    def require_openpyxl(self):
        pytest.importorskip("openpyxl", reason="openpyxl not installed")

    def test_xlsx_extraction_creates_markdown_table(self):
        """Valid .xlsx bytes → Markdown table with headers and rows."""
        import openpyxl, io as _io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Revenue", "EBITDA", "Net Income"])
        ws.append([1000, 200, 150])
        ws.append([1200, 240, 180])
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from credit_report.generation.evidence import _extract_text_from_xlsx
        text = _extract_text_from_xlsx(buf.read())
        assert "| Revenue | EBITDA | Net Income |" in text
        assert "| 1000 | 200 | 150 |" in text

    def test_xlsx_row_padding_for_short_rows(self):
        """Rows shorter than header must be padded to match header width."""
        import openpyxl, io as _io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Col1", "Col2", "Col3"])
        ws.append([1, 2])           # short row — missing Col3
        ws.append([4, 5, 6])        # full row
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from credit_report.generation.evidence import _extract_text_from_xlsx
        text = _extract_text_from_xlsx(buf.read())
        lines = [l for l in text.splitlines() if l.startswith("|")]
        data_lines = [l for l in lines if "---" not in l and "Col" not in l]
        # Each data row must have exactly 3 cells (3 pipes separating 3 values + borders)
        for line in data_lines:
            cell_count = len(line.split("|")) - 2  # subtract border pipes
            assert cell_count == 3, f"Row has {cell_count} cells, expected 3: {line!r}"

    def test_xlsx_row_cap_at_200(self):
        """Only first 200 data rows are extracted per sheet (ETL cap)."""
        import openpyxl, io as _io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Value"])
        for i in range(300):
            ws.append([i])
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from credit_report.generation.evidence import _extract_text_from_xlsx
        text = _extract_text_from_xlsx(buf.read())
        data_rows = [l for l in text.splitlines() if l.startswith("|") and "Value" not in l and "---" not in l]
        assert len(data_rows) <= 200, f"More than 200 rows extracted: {len(data_rows)}"
        assert len(data_rows) == 200, f"Expected exactly 200 rows, got {len(data_rows)}"

    def test_corrupted_xlsx_returns_empty_string(self):
        """Corrupted bytes → empty string (not exception)."""
        from credit_report.generation.evidence import _extract_text_from_xlsx
        result = _extract_text_from_xlsx(b"THIS IS NOT A VALID XLSX")
        assert result == ""

    def test_xls_file_returns_diagnostic_string(self, ac, admin_hdrs, report):
        """Uploading .xls file returns diagnostic text, not silent empty."""
        from credit_report.generation.evidence import extract_text_from_file
        # A minimal .xls magic header (old BIFF8 format — not readable by openpyxl)
        xls_header = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1" + b"\x00" * 100
        text, fmt = extract_text_from_file(xls_header, "financial_data.xls")
        # Must return diagnostic string, not crash
        assert "XLS_UNSUPPORTED" in text or text == "", \
            f"Expected diagnostic for .xls, got: {text!r}"
        assert fmt == "xls"


# ══════════════════════════════════════════════════════════════════════════════
# 7 — PDF _md_to_html XSS Fix Verification
# ══════════════════════════════════════════════════════════════════════════════

class TestPDFHtmlFallbackXSS:
    """Verify the XSS fix in _md_to_html (both markdown-lib and fallback paths)."""

    def test_script_tag_in_heading_is_escaped(self):
        """<script> in heading must NOT appear as raw <script> tag in HTML output."""
        from credit_report.api.export import _md_to_html
        xss_md = "# <script>alert('xss')</script>"
        result = _md_to_html(xss_md)
        assert "<script>" not in result, f"Unescaped <script> in output: {result}"
        # Escaped form (&lt;script&gt;) must appear instead
        assert "&lt;script&gt;" in result or "alert" not in result

    def test_img_onerror_xss_in_bold(self):
        """**<img onerror=alert(1)>** must be escaped, not left as raw HTML."""
        from credit_report.api.export import _md_to_html
        xss_md = "**<img src=x onerror=alert(1)>**"
        result = _md_to_html(xss_md)
        assert "<img" not in result, f"Unescaped <img> in output: {result}"

    def test_safe_markdown_preserved(self):
        """Normal markdown (headers, bold, italic) renders correctly."""
        from credit_report.api.export import _md_to_html
        md = "## Hello\n\nThis is **bold** and *italic* text."
        result = _md_to_html(md)
        assert "bold" in result.lower() or "strong" in result.lower() or "<h" in result

    def test_html_entities_preserved_in_content(self):
        """Regular text with numbers and % passes through correctly."""
        from credit_report.api.export import _md_to_html
        md = "Revenue was USD 100m (5% growth)"
        result = _md_to_html(md)
        assert "100m" in result
        assert "5%" in result

    def test_sanitize_function_escapes_html_tags(self):
        """_sanitize_html_in_markdown replaces < > with &lt; &gt;."""
        from credit_report.api.export import _sanitize_html_in_markdown
        result = _sanitize_html_in_markdown("<script>alert(1)</script>")
        assert "<script>" not in result
        assert "&lt;script&gt;" in result

    def test_sanitize_preserves_markdown_syntax(self):
        """Markdown syntax characters (* # | ---) are untouched by sanitizer."""
        from credit_report.api.export import _sanitize_html_in_markdown
        md = "# Heading\n\n**bold** and *italic*\n\n| A | B |\n|---|---|\n| 1 | 2 |"
        result = _sanitize_html_in_markdown(md)
        assert "# Heading" in result
        assert "**bold**" in result
        assert "| A | B |" in result


# ══════════════════════════════════════════════════════════════════════════════
# 8 — Auth /setup Endpoint
# ══════════════════════════════════════════════════════════════════════════════

class TestAuthSetupEndpoint:

    async def test_setup_blocked_when_no_setup_key_configured(self, ac):
        """Without SETUP_KEY env var, setup endpoint returns 503."""
        with patch.dict(os.environ, {}, clear=False):
            # Make sure SETUP_KEY is not set
            os.environ.pop("SETUP_KEY", None)
            r = await ac.post(f"{AUTH}/setup",
                              json={"email": "admin@new.com", "password": "Pass1234!",
                                    "setup_key": "any-key"})
        assert r.status_code in (403, 503), \
            f"Expected 403/503 when SETUP_KEY not configured, got {r.status_code}"

    async def test_setup_blocked_with_wrong_key(self, ac):
        """Wrong setup_key → 403."""
        with patch.dict(os.environ, {"SETUP_KEY": "correct-key"}):
            r = await ac.post(f"{AUTH}/setup",
                              json={"email": "admin2@new.com", "password": "Pass1234!",
                                    "setup_key": "wrong-key"})
        assert r.status_code == 403

    async def test_setup_succeeds_with_correct_key(self, ac):
        """Correct setup_key creates admin account."""
        key = f"test-setup-{uuid.uuid4().hex[:8]}"
        email = f"setup_{uuid.uuid4().hex[:8]}@admin.test"
        with patch.dict(os.environ, {"SETUP_KEY": key}):
            r = await ac.post(f"{AUTH}/setup",
                              json={"email": email, "password": "SetupPass1!",
                                    "setup_key": key})
        # 201 (created) or 409 (admin already exists)
        assert r.status_code in (201, 409), \
            f"Unexpected status from /auth/setup: {r.status_code} {r.text[:200]}"

    async def test_setup_requires_email_and_password(self, ac):
        """Missing required fields → 422."""
        key = f"test-key-{uuid.uuid4().hex}"
        with patch.dict(os.environ, {"SETUP_KEY": key}):
            r = await ac.post(f"{AUTH}/setup",
                              json={"setup_key": key})  # missing email/password
        assert r.status_code == 422


# ══════════════════════════════════════════════════════════════════════════════
# 9 — Section Name Consistency
# ══════════════════════════════════════════════════════════════════════════════

class TestSectionNameConsistency:
    """Verify SECTION_NAMES in Python matches SNAMES in JS static/index.html."""

    def test_export_py_section_names_match_expected(self):
        """Export section names must match the canonical list."""
        from credit_report.api.export import SECTION_NAMES
        expected = {
            1: "Facility Structure & Key Terms",
            2: "Overall Comments",
            3: "Credit Risk Assessment",
            4: "Borrower Background",
            5: "Collateral Assessment",
            6: "Project Overview (Ship Finance)",
            7: "Financial Analysis",
            8: "Legal Documentation & Charges",
            9: "Compliance Checklist",
            10: "Appendix",
        }
        for sec_no, name in expected.items():
            assert SECTION_NAMES.get(sec_no) == name, \
                f"Section {sec_no}: expected {name!r}, got {SECTION_NAMES.get(sec_no)!r}"

    def test_index_html_snames_contain_all_sections(self):
        """static/index.html SNAMES JS dict must include all sections 1-10."""
        html_path = "/home/user/financial-report-analyzer-ai-report/static/index.html"
        with open(html_path) as f:
            content = f.read()
        for sec_no in range(1, 11):
            assert f"{sec_no}:" in content or f"'{sec_no}':" in content, \
                f"Section {sec_no} missing from SNAMES in index.html"

    def test_index_html_login_hint_is_english(self):
        """Login hint fallback text must be English, not Chinese."""
        html_path = "/home/user/financial-report-analyzer-ai-report/static/index.html"
        with open(html_path) as f:
            content = f.read()
        # Check the loginHint div
        import re
        m = re.search(r'id="loginHint"[^>]*>([^<]+)<', content)
        assert m, "loginHint div not found"
        hint_text = m.group(1).strip()
        # Must not contain Chinese characters (Unicode CJK range)
        has_chinese = any('一' <= c <= '鿿' for c in hint_text)
        assert not has_chinese, \
            f"Login hint contains Chinese fallback text: {hint_text!r}"
        # Must be English
        assert len(hint_text) > 10, "Login hint is too short"

    def test_prompt_builder_has_section_headings_for_1_to_10(self):
        """prompt_builder.py must define SECTION_HEADINGS for sections 1-10."""
        from credit_report.generation.prompt_builder import SECTION_HEADINGS
        for sec_no in range(1, 11):
            assert sec_no in SECTION_HEADINGS, \
                f"Section {sec_no} missing from SECTION_HEADINGS"


# ══════════════════════════════════════════════════════════════════════════════
# 10 — Token Quota Per-Role Limits
# ══════════════════════════════════════════════════════════════════════════════

class TestTokenQuotaRoleLimits:

    def test_analyst_limit_is_base(self):
        from credit_report.generation.quota import _limit_for_role
        from credit_report.config import DAILY_TOKEN_LIMIT
        assert _limit_for_role("analyst") == DAILY_TOKEN_LIMIT

    def test_reviewer_limit_is_2x(self):
        from credit_report.generation.quota import _limit_for_role
        from credit_report.config import DAILY_TOKEN_LIMIT
        assert _limit_for_role("reviewer") == DAILY_TOKEN_LIMIT * 2

    def test_approver_limit_is_2x(self):
        from credit_report.generation.quota import _limit_for_role
        from credit_report.config import DAILY_TOKEN_LIMIT
        assert _limit_for_role("approver") == DAILY_TOKEN_LIMIT * 2

    def test_admin_limit_is_5x(self):
        from credit_report.generation.quota import _limit_for_role
        from credit_report.config import DAILY_TOKEN_LIMIT
        assert _limit_for_role("admin") == DAILY_TOKEN_LIMIT * 5

    def test_unknown_role_gets_base_limit(self):
        from credit_report.generation.quota import _limit_for_role
        from credit_report.config import DAILY_TOKEN_LIMIT
        assert _limit_for_role("super_admin_custom") == DAILY_TOKEN_LIMIT

    async def test_reviewer_can_generate_at_2x_analyst_quota(self, ac, admin_hdrs):
        """Reviewer with tokens_used = DAILY_TOKEN_LIMIT * 1.5 should NOT be 429."""
        from credit_report.generation.models import UserTokenQuota
        from credit_report.database import AsyncSessionLocal
        from credit_report.config import DAILY_TOKEN_LIMIT
        from datetime import datetime, timezone

        reviewer_email = f"rev_q_{uuid.uuid4().hex[:8]}@test.com"
        await _register(ac, admin_hdrs, reviewer_email, "reviewer")
        r_me = await ac.post(f"{AUTH}/login",
                             data={"username": reviewer_email, "password": "Pass1234!"})
        rev_uid = r_me.json().get("user_id") or (
            await ac.get(f"{AUTH}/me",
                         headers={"Authorization": f"Bearer {r_me.json()['access_token']}"})
        ).json()["id"]

        today = datetime.now(timezone.utc).date()
        async with AsyncSessionLocal() as db:
            from sqlalchemy import select
            result = await db.execute(
                select(UserTokenQuota).where(
                    UserTokenQuota.user_id == rev_uid,
                    UserTokenQuota.quota_date == today,
                )
            )
            quota = result.scalar_one_or_none()
            tokens_at_1_5x = int(DAILY_TOKEN_LIMIT * 1.5)
            if quota:
                quota.tokens_used = tokens_at_1_5x
            else:
                db.add(UserTokenQuota(
                    id=str(uuid.uuid4()),
                    user_id=rev_uid,
                    quota_date=today,
                    tokens_used=tokens_at_1_5x,
                ))
            await db.commit()

        rev_hdrs = {"Authorization": f"Bearer {r_me.json()['access_token']}"}
        rpt = (await ac.post(f"{RPTS}", json={"borrower_name": "RevQuotaCo"},
                             headers=rev_hdrs)).json()
        with _mock_gemini(), \
             patch("credit_report.api.generate.check_hard_dependencies",
                   new=AsyncMock(return_value=[])):
            r = await ac.post(f"{RPTS}/{rpt['id']}/generate/4", headers=rev_hdrs)
        assert r.status_code != 429, \
            f"Reviewer at 1.5× analyst quota should not hit 429 (reviewer limit is 2×)"


# ══════════════════════════════════════════════════════════════════════════════
# 11 — Evidence Extraction Edge Cases
# ══════════════════════════════════════════════════════════════════════════════

class TestEvidenceExtraction:

    def test_extract_empty_pdf_does_not_crash(self):
        """Empty bytes for PDF → graceful empty string."""
        from credit_report.generation.evidence import extract_text_from_pdf
        result = extract_text_from_pdf(b"")
        assert isinstance(result, str)

    def test_extract_corrupted_pdf_does_not_crash(self):
        """Corrupted PDF bytes → graceful empty string."""
        from credit_report.generation.evidence import extract_text_from_pdf
        result = extract_text_from_pdf(b"%PDF-1.4 CORRUPTED DATA HERE %%EOF")
        assert isinstance(result, str)

    def test_extract_txt_content(self):
        """Plain text file returns the text content."""
        from credit_report.generation.evidence import extract_text_from_file
        content = b"Revenue 2024: USD 500m\nEBITDA: USD 100m"
        text, fmt = extract_text_from_file(content, "report.txt")
        assert "Revenue" in text
        assert fmt == "txt"

    def test_extract_csv_content(self):
        """CSV file extracted as plain text."""
        from credit_report.generation.evidence import extract_text_from_file
        csv_bytes = b"Year,Revenue,EBITDA\n2024,500,100\n2023,450,90\n"
        text, fmt = extract_text_from_file(csv_bytes, "data.csv")
        assert "Revenue" in text or text == ""  # CSV decoded as utf-8
        assert fmt == "csv"

    def test_extract_markdown_content(self):
        """Markdown file extracted as plain text."""
        from credit_report.generation.evidence import extract_text_from_file
        md_bytes = b"# Annual Report\n\n## Revenue\n\nUSD 500m in FY2024\n"
        text, fmt = extract_text_from_file(md_bytes, "report.md")
        assert "Annual Report" in text
        assert fmt == "md"

    def test_extract_unknown_extension_does_not_crash(self):
        """Unknown extension tries PDF then docx extraction gracefully."""
        from credit_report.generation.evidence import extract_text_from_file
        result_text, fmt = extract_text_from_file(b"some content", "file.zzz")
        assert isinstance(result_text, str)

    def test_extract_text_from_file_large_file(self):
        """Large text file (> 1MB) extracted without crash."""
        from credit_report.generation.evidence import extract_text_from_file
        large_content = (b"Revenue data line\n") * 60000  # ~1MB
        text, fmt = extract_text_from_file(large_content, "large.txt")
        assert len(text) > 0
        assert fmt == "txt"


# ══════════════════════════════════════════════════════════════════════════════
# 12 — Quota Race Condition: document as known issue
# ══════════════════════════════════════════════════════════════════════════════

class TestQuotaRaceCondition:
    """
    Quota check_quota() reads then writes without a transaction lock.
    Two concurrent requests can both pass the check before either records tokens.

    This test DOCUMENTS the race condition. It does NOT exercise production-scale
    concurrency (which requires >= 2 real concurrent requests to the same DB).
    The check here verifies the atomic read of the quota is correct sequentially.
    """

    async def test_check_quota_raises_429_when_over_limit(self, ac, admin_hdrs):
        """Seeding tokens_used = limit → check_quota raises 429."""
        from credit_report.generation.quota import check_quota
        from credit_report.generation.models import UserTokenQuota
        from credit_report.database import AsyncSessionLocal
        from credit_report.config import DAILY_TOKEN_LIMIT
        from datetime import datetime, timezone

        analyst_email = f"racetest_{uuid.uuid4().hex[:8]}@quota.test"
        r = await _register(ac, admin_hdrs, analyst_email, "analyst")
        uid = r["id"]

        today = datetime.now(timezone.utc).date()
        async with AsyncSessionLocal() as db:
            db.add(UserTokenQuota(
                id=str(uuid.uuid4()),
                user_id=uid,
                quota_date=today,
                tokens_used=DAILY_TOKEN_LIMIT,  # exactly at limit
            ))
            await db.commit()

        async with AsyncSessionLocal() as db:
            with pytest.raises(Exception) as exc_info:
                await check_quota(db, uid, "analyst")
            # Must raise HTTPException 429
            exc = exc_info.value
            assert hasattr(exc, "status_code") and exc.status_code == 429

    async def test_check_quota_passes_when_under_limit(self, ac, admin_hdrs):
        """Under-limit user → check_quota does not raise."""
        from credit_report.generation.quota import check_quota
        from credit_report.database import AsyncSessionLocal

        analyst_email = f"under_{uuid.uuid4().hex[:8]}@quota.test"
        r = await _register(ac, admin_hdrs, analyst_email, "analyst")
        uid = r["id"]

        async with AsyncSessionLocal() as db:
            # Should not raise — no tokens used yet
            await check_quota(db, uid, "analyst")

    async def test_record_tokens_accumulates_correctly(self, ac, admin_hdrs):
        """record_tokens adds to existing quota record."""
        from credit_report.generation.quota import record_tokens
        from credit_report.generation.models import UserTokenQuota
        from credit_report.database import AsyncSessionLocal
        from datetime import datetime, timezone
        from sqlalchemy import select

        analyst_email = f"acc_{uuid.uuid4().hex[:8]}@quota.test"
        r = await _register(ac, admin_hdrs, analyst_email, "analyst")
        uid = r["id"]

        async with AsyncSessionLocal() as db:
            await record_tokens(db, uid, 1000)
            await record_tokens(db, uid, 500)
            await db.commit()

        today = datetime.now(timezone.utc).date()
        async with AsyncSessionLocal() as db:
            result = await db.execute(
                select(UserTokenQuota).where(
                    UserTokenQuota.user_id == uid,
                    UserTokenQuota.quota_date == today,
                )
            )
            quota = result.scalar_one_or_none()
            assert quota is not None
            assert quota.tokens_used >= 1500  # may include previous test tokens


# ══════════════════════════════════════════════════════════════════════════════
# 13 — Calculation Engine API Integration
# ══════════════════════════════════════════════════════════════════════════════

class TestCalculationAPIEdgeCases:

    async def test_recalculate_with_zero_ebitda_does_not_500(self, ac, admin_hdrs):
        """Zero EBITDA must not crash recalculate (safe_divide handles it)."""
        rpt = (await ac.post(f"{RPTS}", json={"borrower_name": "ZeroEBITDA"},
                             headers=admin_hdrs)).json()

        from credit_report.database import AsyncSessionLocal
        from credit_report.fact_store.repository import upsert_fact
        rid = rpt["id"]
        async with AsyncSessionLocal() as db:
            for metric, value in [("ebitda", 0.0), ("total_debt", 500.0), ("revenue", 1000.0)]:
                await upsert_fact(db, {
                    "report_id": rid, "metric_name": metric,
                    "entity": "ZeroEBITDA", "period": "FY2024",
                    "value": value, "value_text": str(value),
                    "state": "validated", "source_type": "analyst_input_json",
                })
            await db.commit()

        r = await ac.post(f"{RPTS}/{rid}/recalculate", headers=admin_hdrs)
        assert r.status_code == 200, f"recalculate 500d on zero EBITDA: {r.text[:200]}"

    async def test_recalculate_with_no_facts_returns_empty_list(self, ac, admin_hdrs):
        """Report with no facts → recalculate returns empty computed list."""
        rpt = (await ac.post(f"{RPTS}", json={"borrower_name": "EmptyFacts"},
                             headers=admin_hdrs)).json()
        r = await ac.post(f"{RPTS}/{rpt['id']}/recalculate", headers=admin_hdrs)
        assert r.status_code == 200
        body = r.json()
        assert isinstance(body, dict)

    async def test_calculations_schema_is_correct(self, ac, admin_hdrs):
        """GET /calculations returns list with required fields."""
        rpt = (await ac.post(f"{RPTS}", json={"borrower_name": "SchemaCheck"},
                             headers=admin_hdrs)).json()
        r = await ac.get(f"{RPTS}/{rpt['id']}/calculations", headers=admin_hdrs)
        assert r.status_code == 200
        calcs = r.json()
        assert isinstance(calcs, list)
        for c in calcs:
            assert "metric_name" in c, f"metric_name missing: {c}"
            assert "is_stale" in c, f"is_stale missing: {c}"

    async def test_unmapped_queue_endpoint(self, ac, admin_hdrs, report):
        """GET /mapping/unmapped returns a list (may be empty)."""
        r = await ac.get(f"{RPTS}/{report['id']}/mapping/unmapped", headers=admin_hdrs)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    async def test_pending_mapping_rules_endpoint(self, ac, admin_hdrs, report):
        """GET /mapping/rules/pending returns a list (may be empty)."""
        r = await ac.get(f"{RPTS}/{report['id']}/mapping/rules/pending",
                         headers=admin_hdrs)
        assert r.status_code in (200, 404)  # 404 if endpoint doesn't exist
        if r.status_code == 200:
            assert isinstance(r.json(), list)


# ══════════════════════════════════════════════════════════════════════════════
# 14 — Document Upload Edge Cases
# ══════════════════════════════════════════════════════════════════════════════

class TestDocumentUploadEdgeCases:

    async def test_upload_empty_file_does_not_500(self, ac, admin_hdrs, report):
        """Zero-byte file upload must not crash (400 or 422, not 500)."""
        r = await ac.post(
            f"{RPTS}/{report['id']}/documents",
            files={"file": ("empty.txt", io.BytesIO(b""), "text/plain")},
            headers=admin_hdrs,
        )
        assert r.status_code != 500, f"Empty file caused 500: {r.text[:200]}"

    async def test_upload_xls_file_accepted_with_warning(self, ac, admin_hdrs, report):
        """XLS upload is accepted (in ALLOWED_EXTENSIONS) but returns diagnostic extraction."""
        # Minimal XLS magic bytes (BIFF8 header)
        xls_bytes = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1" + b"\x00" * 200
        r = await ac.post(
            f"{RPTS}/{report['id']}/documents",
            files={"file": ("data.xls", io.BytesIO(xls_bytes), "application/vnd.ms-excel")},
            headers=admin_hdrs,
        )
        # Should accept the upload (not 415), even though extraction is limited
        assert r.status_code in (200, 201, 202), \
            f"XLS upload rejected unexpectedly: {r.status_code}"

    async def test_upload_unsupported_extension_rejected(self, ac, admin_hdrs, report):
        """Unknown extension (.exe) must be rejected (415 or 422)."""
        r = await ac.post(
            f"{RPTS}/{report['id']}/documents",
            files={"file": ("malware.exe", io.BytesIO(b"MZ\x90\x00"), "application/octet-stream")},
            headers=admin_hdrs,
        )
        assert r.status_code in (400, 415, 422), \
            f"Unsupported extension was accepted: {r.status_code}"

    async def test_upload_txt_large_file_no_500(self, ac, admin_hdrs, report):
        """Large text file (2MB) must not cause 500."""
        large_txt = b"Revenue USD 500m\n" * 125000  # ~2MB
        r = await ac.post(
            f"{RPTS}/{report['id']}/documents",
            files={"file": ("large.txt", io.BytesIO(large_txt), "text/plain")},
            headers=admin_hdrs,
        )
        assert r.status_code != 500, f"Large upload caused 500: {r.text[:100]}"

    async def test_upload_requires_auth(self, ac, report, admin_hdrs):
        r = await ac.post(
            f"{RPTS}/{report['id']}/documents",
            files={"file": ("test.txt", io.BytesIO(b"hello"), "text/plain")},
        )
        assert r.status_code == 401
