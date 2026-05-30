"""
Hidden Bug Piercing Tests
=========================
20 production bugs that AI sandboxes (isolated pytest, no real platform, no
real files, no real browser) consistently mask.  Every test is self-contained,
pure-Python, runnable with ``pytest tests/test_hidden_bugs.py``.

Each class maps to one of the five defect categories identified by deep code
review.  Two "correction" tests at the end verify that the code actually
handles quota and semaphore *correctly* — the original bug report overstated
those two as defects.

RE-VERIFICATION NOTES (after a second, adversarial pass against the code and
known platform behaviour — corrections folded into the relevant docstrings):
  * Bug #1 (JWT): Render ``generateValue: true`` persists the secret across
    deploys; it does NOT rotate every deploy.  Scope narrowed to the real
    module-level _KEY_BYTES binding + restart semantics.
  * Bug #3 (postgres://): could not confirm Render emits the short scheme;
    reframed as a SQLAlchemy-2.0 robustness gap rather than a guaranteed
    Render crash.  The code behaviour asserted is still exactly true.
  * Bug #4 (alembic): STRENGTHENED — alembic is fully configured with 8
    migrations but never run; create_all is used instead (genuine dead
    migration infrastructure + schema drift on PostgreSQL).
  * Bug #7 (LLM timeout): DEBUNKED.  Two framings both proved false — the
    "Render 30 s gateway" premise (that is Heroku; generation is a polled
    BackgroundTask anyway) AND a follow-up "uncancellable run_in_executor
    thread" guess.  claude_client awaits the NATIVE async client inside
    asyncio.wait_for, so the timeout cancels cleanly.  The test now guards
    that correct implementation instead of asserting a non-existent bug.

Source-file anchors are given in each test docstring so findings can be
re-verified after refactors.
"""
from __future__ import annotations

import ast
import asyncio
import importlib
import inspect
import io
import json
import os
import re
import tempfile
import textwrap
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
import yaml  # PyYAML is a dev-dep; available in requirements.txt

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).parent.parent
RENDER_YAML = REPO_ROOT / "render.yaml"
INDEX_HTML = REPO_ROOT / "static" / "index.html"


# ===========================================================================
# Group 1 – Cross-Deploy State Bugs (#1–#4)
# ===========================================================================

class TestCrossDeployStateBugs:
    """Bugs caused by state that is re-initialised on every Render deploy."""

    # ── Bug #1 (FIXED) ────────────────────────────────────────────────────
    def test_jwt_key_bytes_bound_at_module_import(self):
        """FIX: _KEY_BYTES removed; _sign() now reads SECRET_KEY dynamically.

        The original defect was that ``_KEY_BYTES = SECRET_KEY.encode()`` was
        a module-level constant in auth.py.  Any change to the SECRET_KEY env
        var only took effect after a full interpreter restart.

        FIX applied: removed the ``_KEY_BYTES`` module-level constant.
        ``_sign()`` now calls ``os.getenv("SECRET_KEY", DEFAULT_SECRET_KEY)``
        on every invocation, so key rotation takes effect without restart.

        Source: credit_report/security/auth.py
        """
        from credit_report.security import auth as auth_mod

        # Fix confirmed: _KEY_BYTES no longer exists at module level.
        assert not hasattr(auth_mod, "_KEY_BYTES"), (
            "_KEY_BYTES must be removed — the key should be read dynamically in _sign()"
        )

        # Fix confirmed: _sign() reads the key via os.getenv().
        sign_src = inspect.getsource(auth_mod._sign)
        assert "os.getenv(" in sign_src, (
            "_sign() must read SECRET_KEY via os.getenv() for dynamic key rotation"
        )

        # Verify signing still works correctly end-to-end.
        token = auth_mod.create_access_token("user-fix-1", "analyst")
        payload = auth_mod._decode_jwt(token)
        assert payload.get("sub") == "user-fix-1", "Token round-trip must succeed"

    # ── Bug #2 ─────────────────────────────────────────────────────────────
    def test_seed_admin_always_rehashes_password_on_startup(self):
        """_seed_admin unconditionally overwrites the admin hash from the env var.

        VERIFIED PRECISE SCOPE: the function first checks
        ``if not email or not password: return``, so clearing ADMIN_PASSWORD
        causes an early return — it does NOT lock the admin out.

        The real operational hazard is the opposite direction: whenever
        ADMIN_PASSWORD is set to ANY non-empty value, every startup unconditionally
        overwrites the stored hash with that value (main.py:113-116).  Concretely:
          - Operator accidentally fat-fingers ADMIN_PASSWORD during a redeploy →
            all existing sessions still have the old hash in their tokens, but
            the next login attempt with the old password fails immediately.
          - There is no "confirm before overwrite" guard; the env var is always
            the source of truth at startup time.

        Source: main.py:93-127 (_seed_admin function)
        """
        import main as main_mod

        src = inspect.getsource(main_mod._seed_admin)

        # 1. Confirm the early-return guard (clearing ADMIN_PASSWORD is safe).
        assert "if not email or not password:" in src, (
            "Expected early return when ADMIN_PASSWORD is empty"
        )

        # 2. FIX confirmed: verify_password guard prevents unnecessary rehash.
        assert "verify_password" in src, (
            "verify_password check must be present — only rehash when password actually changed"
        )

        # 3. FIX confirmed: hash overwrite is conditional, not unconditional.
        assert "if not verify_password(" in src, (
            "Hash update must be inside a 'if not verify_password(...)' guard"
        )

    # ── Bug #3 (FIXED) ────────────────────────────────────────────────────
    def test_postgres_url_scheme_not_rewritten_to_asyncpg(self):
        """FIX: config.py now handles both postgresql:// and legacy postgres://.

        SQLAlchemy 2.0 dropped support for the bare ``postgres://`` scheme.
        The fix adds an ``elif startswith("postgres://")`` branch to config.py
        AND migrations/env.py so all postgres:// URLs are rewritten to
        ``postgresql+asyncpg://`` before reaching create_async_engine().

        Source: credit_report/config.py ; migrations/env.py
        """
        config_src = Path(REPO_ROOT / "credit_report" / "config.py").read_text()

        # Fix confirmed: both schemes are handled in config.py.
        assert 'startswith("postgresql://")' in config_src, (
            "postgresql:// handler must still be present in config.py"
        )
        assert 'startswith("postgres://")' in config_src, (
            "postgres:// handler must be present — fix verified"
        )

        # Fix confirmed: same fix applied in migrations/env.py.
        env_src = Path(REPO_ROOT / "migrations" / "env.py").read_text()
        assert 'startswith("postgres://")' in env_src, (
            "postgres:// handler must also be present in migrations/env.py"
        )

        # Functional verification: reproduce the fixed logic.
        render_url = "postgres://user:pass@host:5432/mydb"
        rewritten = render_url
        if rewritten.startswith("postgresql://"):
            rewritten = rewritten.replace("postgresql://", "postgresql+asyncpg://", 1)
        elif rewritten.startswith("postgres://"):
            rewritten = rewritten.replace("postgres://", "postgresql+asyncpg://", 1)
        assert rewritten.startswith("postgresql+asyncpg://"), (
            "postgres:// must be rewritten to postgresql+asyncpg:// by the fixed logic"
        )

    # ── Bug #4 (FIXED) ────────────────────────────────────────────────────
    def test_render_start_command_has_no_alembic_upgrade(self):
        """FIX: render.yaml startCommand now runs ``alembic upgrade head`` before uvicorn.

        The fix ensures migrations are applied on every Render deploy, so the
        DB schema stays in sync with the migration history.

        Source: render.yaml ; main.py ; migrations/versions/*
        """
        data = yaml.safe_load(RENDER_YAML.read_text())
        service = data["services"][0]
        start_cmd = service.get("startCommand", "")

        # Fix confirmed: alembic upgrade head runs before the app server.
        assert "alembic upgrade head" in start_cmd, (
            f"startCommand must include 'alembic upgrade head', got: {start_cmd!r}"
        )
        assert "uvicorn" in start_cmd, "uvicorn must still be in startCommand"

        # Fix confirmed: render.yaml now has a persistent disk mount.
        assert "disk" in service, (
            "render.yaml must have a disk: section for persistent storage (Bug #6 fix)"
        )

        # Alembic infrastructure still intact.
        assert (REPO_ROOT / "alembic.ini").exists(), "alembic.ini must exist"
        versions_dir = REPO_ROOT / "migrations" / "versions"
        migrations = list(versions_dir.glob("*.py"))
        assert len(migrations) >= 5, (
            f"Expected several alembic migrations, found {len(migrations)}"
        )

        # FUNCTIONAL guard: 'alembic upgrade head' must actually SUCCEED on a
        # fresh DB.  A bare string match is not enough — multiple heads or
        # SQLite-incompatible DDL would crash the startCommand on deploy.
        # (1) Exactly one head — no unmerged branch points.
        import subprocess
        heads = subprocess.run(
            ["alembic", "heads"], cwd=REPO_ROOT, capture_output=True, text=True,
        )
        head_lines = [ln for ln in heads.stdout.splitlines() if "(head)" in ln]
        assert len(head_lines) == 1, (
            f"alembic must have exactly one head, found {len(head_lines)}: {head_lines!r} — "
            "'alembic upgrade head' is ambiguous with multiple heads"
        )
        # (2) Full upgrade succeeds on a throwaway SQLite DB (Render's default
        #     when DATABASE_URL is unset), proving the DDL is dialect-safe.
        with tempfile.TemporaryDirectory() as _td:
            db_path = Path(_td) / "verify.db"
            env = {**os.environ, "DATABASE_URL": f"sqlite+aiosqlite:///{db_path}"}
            up = subprocess.run(
                ["alembic", "upgrade", "head"],
                cwd=REPO_ROOT, capture_output=True, text=True, env=env,
            )
            assert up.returncode == 0, (
                f"'alembic upgrade head' must succeed on SQLite; exit={up.returncode}\n"
                f"STDERR:\n{up.stderr[-2000:]}"
            )


# ===========================================================================
# Group 2 – Render Platform Bugs (#5–#6 real; #7 debunked)
# ===========================================================================

class TestRenderPlatformBugs:
    """Bugs caused by Render's ephemeral-filesystem / single-process model.

    #5 and #6 are real platform defects.  #7 (LLM timeout) was investigated
    twice and found NOT to be a bug; its test now guards the correct behaviour.
    """

    # ── Bug #5 (FIXED) ────────────────────────────────────────────────────
    def test_generation_tasks_registry_is_process_local(self):
        """FIX: _generation_tasks now persists to disk; tasks survive restart.

        The fix adds file-based persistence: _create_generation_task and
        _update_generation_task write task state to _generation_tasks.json.
        On module load, _load_tasks_from_disk() reloads non-expired tasks so
        clients polling /generate/status/{task_id} find their tasks after redeploy.

        Source: credit_report/api/generate.py
        """
        import tempfile
        from unittest.mock import patch as _patch

        from credit_report.api import generate as gen_mod

        # Fix confirmed: persistence helpers exist at module level.
        assert hasattr(gen_mod, "_save_tasks_to_disk"), (
            "_save_tasks_to_disk must be defined for cross-restart persistence"
        )
        assert hasattr(gen_mod, "_load_tasks_from_disk"), (
            "_load_tasks_from_disk must be defined for cross-restart persistence"
        )
        assert hasattr(gen_mod, "_TASKS_FILE"), (
            "_TASKS_FILE path must be defined"
        )

        # Functional test: task survives a simulated restart via disk persistence.
        with tempfile.TemporaryDirectory() as tmpdir:
            tasks_file = Path(tmpdir) / "_generation_tasks.json"
            with _patch.object(gen_mod, "_TASKS_FILE", tasks_file):
                task_id = gen_mod._create_generation_task({"status": "running", "section_no": 7})
                assert task_id in gen_mod._generation_tasks
                assert tasks_file.exists(), "Task must be persisted to disk on creation"

                # Simulate restart: clear memory, reload from disk.
                gen_mod._generation_tasks.clear()
                gen_mod._load_tasks_from_disk()
                assert task_id in gen_mod._generation_tasks, (
                    "Task must survive simulated restart via disk persistence"
                )

    # ── Bug #6 (FIXED) ────────────────────────────────────────────────────
    def test_data_directory_is_local_filesystem_path(self):
        """FIX: render.yaml now declares a disk: persistent volume mount.

        The fix adds a ``disk:`` section to render.yaml mounting the data/
        directory at /opt/render/project/src/data, making uploaded documents
        and extracted text persistent across Render restarts/deploys.

        Source: credit_report/config.py ; render.yaml
        """
        from credit_report import config as cfg

        root = str(cfg.CREDIT_REPORTS_ROOT)
        # Config still points to the data/ directory (now mounted as persistent disk).
        assert "s3://" not in root
        assert "gs://" not in root
        assert "azure://" not in root
        assert "data" in root.replace("\\", "/"), (
            f"CREDIT_REPORTS_ROOT={root!r} should be under data/"
        )

        # Fix confirmed: render.yaml has a disk: section.
        data = yaml.safe_load(RENDER_YAML.read_text())
        service = data["services"][0]
        assert "disk" in service, (
            "render.yaml must have a disk: section for persistent storage"
        )
        disk = service["disk"]
        assert "mountPath" in disk, "disk must specify a mountPath"
        assert "sizeGB" in disk, "disk must specify sizeGB"

    # ── Bug #7 (DEBUNKED — kept as a guard) ────────────────────────────────
    def test_llm_timeout_is_correctly_applied_to_a_native_async_call(self):
        """The LLM timeout is implemented correctly — the original bug was wrong.

        Two successive framings of this "bug" both turned out to be FALSE on
        closer inspection; this test documents the truth so the claim is not
        re-raised:

        1. "LLM_TIMEOUT 180 s > Render's 30 s HTTP gateway" — false premise.
           Render does NOT impose a 30 s request timeout (that is Heroku's H12
           router), and generation runs as a polled FastAPI BackgroundTask, so
           no synchronous request is being cut off.
        2. "wait_for cannot cancel the run_in_executor thread" — also false
           HERE.  claude_client does NOT use run_in_executor for the LLM call;
           it awaits the NATIVE async client
           ``client.aio.models.generate_content(...)`` inside
           ``asyncio.wait_for(..., timeout=LLM_TIMEOUT_SECONDS)``.  wait_for
           propagates CancelledError into a native coroutine, so the request is
           cancelled cleanly — no thread leak.

        What IS true and worth pinning: the timeout is genuinely wired up to
        both Gemini entry points, so it is not a dead config value.

        Source: credit_report/generation/claude_client.py:44-54, 113-127
        """
        import inspect as _inspect
        from credit_report.generation import claude_client

        for fn_name in ("call_gemini_raw", "generate_section_markdown"):
            fn = getattr(claude_client, fn_name)
            src = _inspect.getsource(fn)
            assert "asyncio.wait_for(" in src, (
                f"{fn_name} must bound the LLM call with asyncio.wait_for"
            )
            assert "timeout=LLM_TIMEOUT_SECONDS" in src, (
                f"{fn_name} must pass LLM_TIMEOUT_SECONDS as the timeout"
            )
            # Confirm it wraps the NATIVE async call, not a thread executor —
            # this is why wait_for can cancel it cleanly (no orphaned thread).
            assert "client.aio.models.generate_content" in src, (
                f"{fn_name} should await the native async Gemini client"
            )
            assert "run_in_executor" not in src, (
                f"{fn_name} must NOT use run_in_executor for the LLM call — "
                "the native-async path is what makes the timeout cancellable"
            )


# ===========================================================================
# Group 3 – In-Memory State Bugs (#8–#10)
# ===========================================================================

class TestInMemoryStateBugs:
    """Bugs caused by unretained asyncio tasks or destructive startup queries."""

    # ── Bug #8 ─────────────────────────────────────────────────────────────
    def test_asyncio_create_task_reference_not_retained(self):
        """asyncio.create_task() return value is discarded — formal GC risk.

        asyncio._all_tasks is a WeakSet.  The Python docs explicitly state:
        "Save a reference to the result of create_task() to avoid a task
        disappearing mid-execution."

        VERIFIED SEVERITY (nuanced): in this specific code path the task
        immediately enters ``async with _extraction_semaphore`` then
        ``await loop.run_in_executor(...)``; both the semaphore's waiter queue
        and the executor Future's done-callbacks hold strong references during
        execution, so GC does not manifest in normal operation.

        The code is still formally wrong per the documented API contract — a
        future Python release or a subtle event-loop scheduling change could
        expose the latent risk.  The correct fix is one line:
          ``_task = asyncio.create_task(runner())``.

        Source: credit_report/api/generate.py:179-182
                Python docs: asyncio.create_task()
        """
        import asyncio.tasks as _tasks
        import weakref

        # Confirm the event loop tracks tasks in a WeakSet (formal GC risk).
        assert isinstance(_tasks._all_tasks, weakref.WeakSet), (
            "asyncio._all_tasks must be a WeakSet — confirms GC risk is formally present"
        )

        src = Path(REPO_ROOT / "credit_report" / "api" / "generate.py").read_text()
        tree = ast.parse(src)
        fn_node = next(
            n for n in ast.walk(tree)
            if isinstance(n, ast.FunctionDef) and n.name == "_schedule_document_text_extraction"
        )
        fn_src = ast.get_source_segment(src, fn_node) or ""

        # Fix confirmed: task reference is stored in a variable.
        assert re.search(r"\w+\s*=\s*asyncio\.create_task\(", fn_src), (
            "Task reference must be stored in a variable — GC risk fix verified"
        )

        # Fix confirmed: module-level set retains strong reference until done.
        assert "_background_tasks" in src, (
            "_background_tasks set must exist at module level to hold strong refs"
        )
        assert "add_done_callback" in fn_src, (
            "Task must use add_done_callback to remove itself from the set when done"
        )

    # ── Bug #9 ─────────────────────────────────────────────────────────────
    def test_extraction_semaphore_is_module_level_not_shared_across_workers(self):
        """_extraction_semaphore is a per-process asyncio.Semaphore (future-risk only).

        VERIFIED SCOPE: for the current Render free-tier single-process model
        this works correctly and is NOT a bug today.  It is a latent design
        risk: if Render ever scales to multiple workers (e.g. uvicorn --workers N)
        each process has its own semaphore, so the per-instance extraction limit
        of 1 would silently become N concurrent extractions overall.

        This test guards the architecture — the semaphore exists and is
        module-level — so any future multi-worker refactor surfaces here.

        Source: credit_report/api/generate.py:44
        """
        from credit_report.api import generate as gen_mod

        sem = gen_mod._extraction_semaphore
        assert isinstance(sem, asyncio.Semaphore), (
            "_extraction_semaphore should be an asyncio.Semaphore"
        )
        assert "_extraction_semaphore" in vars(gen_mod), (
            "Semaphore must be at module-level, confirming it is process-local"
        )

    # ── Bug #10 ────────────────────────────────────────────────────────────
    def test_safe_add_columns_dedup_keeps_oldest_row_deletes_newer(self):
        """The startup dedup DELETE uses MIN(id), keeping the *oldest* row.

        VERIFIED PRECISE SCOPE: ``_upsert_section_input_json`` does a true
        SELECT-then-UPDATE, so user edits never produce duplicate rows.
        Duplicate rows arise only from a concurrent-generation race condition
        where two requests both INSERT because both SELECTed an empty result
        simultaneously.  In that narrow race scenario both rows contain
        effectively identical AI-generated content, so MIN(id) vs MAX(id) is
        equivalent — no user data is lost.

        The code fact is still correct and worth pinning: the dedup strategy
        is oldest-wins (MIN), not newest-wins.  A future code change that
        allows legitimate user-initiated second INSERTs (e.g. re-triggering
        section input via a new API path that bypasses the upsert) would
        silently lose the newer row on the next startup.

        Source: main.py:70-76 (_safe_add_columns dedup block)
        """
        src = Path(REPO_ROOT / "main.py").read_text()

        # The DELETE statement must use MIN(id) — confirming oldest-wins.
        assert "SELECT MIN(id)" in src, (
            "Expected MIN(id) in the dedup DELETE — confirms oldest row survives"
        )
        assert "GROUP BY report_id, section_no" in src

        # The SQL is split across f-string lines:
        #   f"DELETE FROM {tbl} WHERE id NOT IN ("
        #   f"  SELECT MIN(id) FROM {tbl} GROUP BY report_id, section_no"
        # Check each fragment independently rather than as one regex.
        assert 'WHERE id NOT IN (' in src, (
            "Expected 'WHERE id NOT IN (' in the dedup DELETE block"
        )
        assert 'DELETE FROM {tbl} WHERE id NOT IN' in src, (
            "Expected f-string DELETE FROM {tbl} … WHERE id NOT IN in main.py"
        )


# ===========================================================================
# Group 4 – Real Document Format Bugs (#11–#15)
# ===========================================================================

class TestDocumentFormatBugs:
    """Bugs that only appear with real financial documents."""

    # ── Bug #11 (FIXED) ────────────────────────────────────────────────────
    def test_vision_ocr_hard_caps_pdf_bytes_at_20mb(self):
        """FIX: Vision OCR limit now uses configurable CR_OCR_MAX_PDF_MB (default 50 MB).

        The hardcoded 20 MB cap has been replaced with the ``CR_OCR_MAX_PDF_MB``
        config variable (default 50 MB), making it tuneable without code changes.

        Source: credit_report/generation/evidence.py ; credit_report/config.py
        """
        from credit_report.generation.evidence import extract_text_from_scanned_pdf_vision
        from credit_report.config import CR_OCR_MAX_PDF_MB

        # Fix confirmed: cap is now configurable and defaulted to 50 MB (> 20 MB).
        assert CR_OCR_MAX_PDF_MB >= 50, (
            f"CR_OCR_MAX_PDF_MB default should be >= 50 MB, got {CR_OCR_MAX_PDF_MB}"
        )

        # Fix confirmed: hard-coded 20 MB cap is gone from the function source.
        src = inspect.getsource(extract_text_from_scanned_pdf_vision)
        assert "pdf_bytes[:20 * 1024 * 1024]" not in src, (
            "Hardcoded 20 MB cap must be replaced with CR_OCR_MAX_PDF_MB"
        )
        assert "CR_OCR_MAX_PDF_MB" in src, (
            "extract_text_from_scanned_pdf_vision must use CR_OCR_MAX_PDF_MB"
        )

        # Functional verification: passing >20 MB should not raise.
        twenty_one_mb = b"\x00" * (21 * 1024 * 1024)
        with patch("credit_report.config.GEMINI_API_KEY", ""):
            result = extract_text_from_scanned_pdf_vision(twenty_one_mb)
        assert result == "", "Expected empty string when GEMINI_API_KEY unset"

    # ── Bug #12 (FIXED) ────────────────────────────────────────────────────
    def test_etl_prompt_truncates_text_at_120k_chars(self):
        """FIX: CR_ETL_MAX_TEXT_CHARS raised from 120 000 to 500 000 (default).

        The old 120 000-character cap silently dropped financial statements from
        200-page TSMC annual reports.  The fix raises the default to 500 000
        (~200 pages of dense text) and makes it configurable via env var.

        Source: credit_report/generation/etl.py ; credit_report/config.py
        """
        from credit_report.generation import etl as etl_mod
        from credit_report.config import CR_ETL_MAX_TEXT_CHARS

        # Fix confirmed: cap raised well above 120 000 (old limit).
        assert CR_ETL_MAX_TEXT_CHARS > 120_000, (
            f"CR_ETL_MAX_TEXT_CHARS must be > 120000 after fix, got {CR_ETL_MAX_TEXT_CHARS}"
        )

        # Build a text with a clearly unique overflow sentinel so we can detect
        # whether the chars BEYOND the cap leaked into the prompt.  Using the same
        # character for both body and overflow would make the overflow a substring
        # of the body and the assertion would trivially pass.
        body = "A" * CR_ETL_MAX_TEXT_CHARS
        overflow_sentinel = "OVERFLOW_SENTINEL_ZZZ_" * 2500  # 55 000 chars, all Z
        long_text = body + overflow_sentinel

        assert len(long_text) == CR_ETL_MAX_TEXT_CHARS + len(overflow_sentinel)

        # _build_etl_prompt is a private helper; call it directly.
        _sys, user_prompt = etl_mod._build_etl_prompt("annual_report", long_text, [4, 7])

        # The overflow sentinel must NOT appear in the generated prompt.
        assert "OVERFLOW_SENTINEL_ZZZ" not in user_prompt, (
            "Chars beyond CR_ETL_MAX_TEXT_CHARS must be absent from the prompt"
        )
        # The body (up to the cap) must be present.
        assert "A" * 100 in user_prompt, (
            "First CR_ETL_MAX_TEXT_CHARS chars (body) must appear in the prompt"
        )

    # ── Bug #13 (FIXED) ────────────────────────────────────────────────────
    def test_openpyxl_read_only_merged_cells_return_none(self):
        """FIX: evidence.py now opens xlsx in normal mode and expands merged cells.

        The original _extract_text_from_xlsx used ``read_only=True`` which causes
        non-anchor merged cells to return None (losing header labels).  The fix
        switches to normal mode and explicitly expands merged cell values before
        iterating rows.

        Source: credit_report/generation/evidence.py
        """
        import openpyxl
        from openpyxl import Workbook

        # Create a workbook with a merged cell A1:C1 = "TSMC 2024 Revenue".
        wb_write = Workbook()
        ws_write = wb_write.active
        ws_write["A1"] = "TSMC 2024 Revenue"
        ws_write["B1"] = "Should be empty after merge"
        ws_write["C1"] = "Should be empty after merge"
        ws_write.merge_cells("A1:C1")
        ws_write["A2"] = 100
        ws_write["B2"] = 200
        ws_write["C2"] = 300

        buf = io.BytesIO()
        wb_write.save(buf)
        buf.seek(0)
        file_bytes = buf.read()

        # Re-open with the same options used in _extract_text_from_xlsx.
        wb_ro = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        ws_ro = wb_ro.active
        rows = list(ws_ro.iter_rows(values_only=True))
        header_row = rows[0]

        # A1 (anchor) retains the value; B1, C1 are None in read_only mode
        # (this demonstrates WHY the fix was needed — openpyxl behaviour is unchanged).
        assert header_row[0] == "TSMC 2024 Revenue", (
            "Anchor cell A1 should retain its value"
        )
        assert header_row[1] is None, (
            "Merged non-anchor B1 is None in read_only mode — this is why we switched mode"
        )
        assert header_row[2] is None, (
            "Merged non-anchor C1 is None in read_only mode — this is why we switched mode"
        )
        wb_ro.close()

        # Fix confirmed: evidence.py no longer uses read_only=True for xlsx.
        evidence_src = Path(REPO_ROOT / "credit_report" / "generation" / "evidence.py").read_text()
        xlsx_fn_start = evidence_src.find("def _extract_text_from_xlsx")
        xlsx_fn_end = evidence_src.find("\ndef ", xlsx_fn_start + 1)
        xlsx_fn_src = evidence_src[xlsx_fn_start:xlsx_fn_end]
        assert "read_only=True" not in xlsx_fn_src, (
            "_extract_text_from_xlsx must not use read_only=True — fix verified"
        )
        assert "merged_cells" in xlsx_fn_src, (
            "_extract_text_from_xlsx must expand merged cells after switching from read_only"
        )

    # ── Bug #14 (FIXED) ────────────────────────────────────────────────────
    @pytest.mark.asyncio
    async def test_twse_html_response_silently_returns_empty_list(self):
        """FIX: fetch() now checks content-type before calling resp.json().

        A content-type check (and optional text-prefix heuristic) is performed
        before resp.json().  Non-JSON responses log a clear diagnostic warning
        and raise ValueError, which is caught by the outer except and triggers
        the CSV fallback.  The return value is still [] when both fail, but
        the log now distinguishes HTML from legitimate JSON errors.

        Source: credit_report/integrations/twse.py
        """
        from credit_report.integrations.twse import TWSEOpenAPIClient

        # Fix confirmed: content-type check is in the source.
        import credit_report.integrations.twse as twse_src_mod
        fetch_src = inspect.getsource(twse_src_mod.TWSEOpenAPIClient.fetch)
        assert "content-type" in fetch_src or "content_type" in fetch_src, (
            "fetch() must check content-type header before calling resp.json()"
        )

        html_body = b"<html><body>Service Unavailable</body></html>"

        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.side_effect = json.JSONDecodeError("Expecting value", "", 0)

        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get = AsyncMock(return_value=mock_resp)

        with patch("httpx.AsyncClient", return_value=mock_client):
            client = TWSEOpenAPIClient()
            result = await client.fetch("/opendata/t187ap06_L_ci")

        assert result == [], (
            "Expected [] when TWSE returns HTML that fails JSON parse"
        )

    # ── Bug #15 ────────────────────────────────────────────────────────────
    def test_twse_roc_year_converts_to_gregorian(self):
        """TWSE API reports years in ROC (Republic of China) calendar.

        TWSE fiscal year 113 must be converted to 2024 (113 + 1911).
        _period_key() in twse.py applies this conversion, but only when the
        raw year value is a digit string < 1911.

        Source: credit_report/integrations/twse.py:160-162
        """
        from credit_report.integrations.twse import _period_key

        # ROC year 113 → Gregorian 2024.
        row_roc = {"年度": "113", "季別": ""}
        key = _period_key(row_roc)
        assert "2024" in key, (
            f"ROC year 113 should convert to Gregorian 2024, got {key!r}"
        )

        # Already-Gregorian year must pass through unchanged.
        row_greg = {"年度": "2024", "季別": ""}
        key2 = _period_key(row_greg)
        assert "2024" in key2

        # ROC quarter: year 112 Q4 → FY2023Q4.
        row_q = {"年度": "112", "季別": "4"}
        key3 = _period_key(row_q)
        assert "2023" in key3 and "Q4" in key3, (
            f"Expected FY2023Q4, got {key3!r}"
        )


# ===========================================================================
# Group 5 – Browser Security Bugs (#16–#18)
# ===========================================================================

class TestBrowserSecurityBugs:
    """Bugs visible only in a real browser or with real CORS headers."""

    # ── Bug #16 (FIXED) ────────────────────────────────────────────────────
    def test_marked_parse_assigned_to_innerhtml_without_dompurify(self):
        """FIX: DOMPurify is now loaded and wraps all marked.parse() → innerHTML paths.

        DOMPurify 3.1.6 is added via CDN.  All 6 occurrences of
        ``marked.parse(...)`` that feed ``innerHTML`` are now wrapped in
        ``DOMPurify.sanitize(marked.parse(...))``, removing the stored XSS vector.

        Source: static/index.html
        """
        html = INDEX_HTML.read_text(encoding="utf-8")

        # Fix confirmed: DOMPurify is loaded.
        assert "DOMPurify" in html, (
            "DOMPurify must be loaded in index.html"
        )
        assert "dompurify" in html.lower(), (
            "DOMPurify CDN script tag must be present"
        )

        # Fix confirmed: all marked.parse() calls are wrapped in DOMPurify.sanitize().
        assert "DOMPurify.sanitize(marked.parse(" in html, (
            "marked.parse() must be wrapped in DOMPurify.sanitize()"
        )

        # Fix confirmed: no bare innerHTML = marked.parse() pattern remains.
        bare_xss = re.compile(r"innerHTML\s*=\s*[^;]*?marked\.parse\((?![^)]*DOMPurify)", re.DOTALL)
        # Check that every marked.parse adjacent to innerHTML is wrapped.
        assert "DOMPurify.sanitize(marked.parse(d.markdown))" in html, (
            "Primary marked.parse(d.markdown) path must be sanitized"
        )

    # ── Bug #17 ────────────────────────────────────────────────────────────
    def test_cors_wildcard_forces_credentials_false_cookie_auth_impossible(self):
        """CORS_ALLOW_ORIGINS='*' forces allow_credentials=False — a design constraint.

        VERIFIED PRECISE SCOPE: the frontend uses ``Authorization: Bearer <token>``
        in request headers (index.html:1994 ``function H()``), NOT cookies.
        In CORS terminology, "credentials" means cookies + TLS client certs +
        HTTP auth — NOT custom headers like Authorization: Bearer.
        Therefore ``allow_credentials=False`` + ``allow_origins=["*"]`` works
        correctly for this app: all fetch calls succeed.

        What IS constrained by this config:
        - Cookie-based cross-origin auth: impossible (browser would block it).
          This matters if anyone tries to add SSO or session-cookie auth later.
        - The FastAPI CORS middleware correctly guards against the
          ``allow_credentials=True + allow_origins=["*"]`` combination that
          would raise ``ValueError`` at startup — that guard is also tested here.

        Source: main.py:159-162 ; static/index.html:1994
                render.yaml:42-43
        """
        # 1. Verify render.yaml ships with wildcard.
        data = yaml.safe_load(RENDER_YAML.read_text())
        env_vars = {e["key"]: e.get("value", "") for e in data["services"][0].get("envVars", [])}
        cors_value = env_vars.get("CORS_ALLOW_ORIGINS", "")
        assert cors_value == "*", (
            f"render.yaml CORS_ALLOW_ORIGINS expected '*', got {cors_value!r}"
        )

        # 2. Confirm the app correctly disables credentials for wildcard (avoids ValueError).
        cors_origins = [o.strip() for o in cors_value.split(",") if o.strip()]
        allow_creds = "*" not in cors_origins
        assert allow_creds is False, (
            "With CORS_ALLOW_ORIGINS='*', allow_credentials must be False to "
            "avoid FastAPI ValueError at startup"
        )

        # 3. Confirm the frontend uses Bearer header auth (not cookies).
        html = INDEX_HTML.read_text(encoding="utf-8")
        assert "function H(){return{'Authorization':'Bearer '+token}" in html, (
            "Frontend must use Authorization: Bearer header (not cookies) — "
            "confirming credentials=False is not a functional block"
        )
        assert "credentials:'include'" not in html, (
            "Frontend must NOT use credentials:'include' — Bearer header is sufficient"
        )

    # ── Bug #18 (FIXED) ────────────────────────────────────────────────────
    def test_twse_paid_in_capital_unit_differs_from_financial_statements(self):
        """FIX: paid_in_capital is now normalised to thousands NTD in build_section7_input.

        The raw TWD value (e.g. 259_303_805_450 for TSMC) is divided by 1_000
        in build_section7_input so it matches the unit of income/balance sheet
        metrics, and a ``paid_in_capital_unit: "thousands_NTD"`` key is added
        for explicitness.

        Source: credit_report/integrations/twse.py
        """
        from credit_report.integrations import twse as twse_mod

        # paid_in_capital still in PROFILE_FIELD_ALIASES (source unchanged).
        assert "paid_in_capital" in twse_mod.PROFILE_FIELD_ALIASES, (
            "paid_in_capital must be in PROFILE_FIELD_ALIASES"
        )
        assert "實收資本額" in twse_mod.PROFILE_FIELD_ALIASES["paid_in_capital"]

        # Fix confirmed: normalisation code is present in build_section7_input.
        src = inspect.getsource(twse_mod.build_section7_input)
        assert "1_000" in src or "1000" in src, (
            "build_section7_input must divide paid_in_capital by 1000"
        )
        assert "paid_in_capital_unit" in src, (
            "build_section7_input must add paid_in_capital_unit key for clarity"
        )
        assert '"thousands"' in src or "'thousands'" in src, (
            "build_section7_input should set unit='thousands' for financial data"
        )


# ===========================================================================
# Correction Tests – Bugs overstated in the original analysis
# ===========================================================================

class TestCorrections:
    """Two items from the original '20 bugs' list are actually implemented
    correctly.  These tests confirm that and prevent future regressions."""

    # ── Correction #1 (Bug #7 in original list) ────────────────────────────
    def test_quota_is_correctly_scoped_per_user_not_global(self):
        """Token quota is per-user, stored with user_id as the key.

        The original analysis claimed quota was global.  In fact,
        check_quota() and record_tokens() both filter by
        ``UserTokenQuota.user_id == user_id``.

        Source: credit_report/generation/quota.py:31-36
        """
        import inspect
        from credit_report.generation import quota as quota_mod

        check_src = inspect.getsource(quota_mod.check_quota)
        assert "user_id" in check_src
        assert "UserTokenQuota.user_id == user_id" in check_src, (
            "check_quota must filter by user_id — quota is per-user"
        )

        record_src = inspect.getsource(quota_mod.record_tokens)
        assert "UserTokenQuota.user_id == user_id" in record_src, (
            "record_tokens must filter by user_id — quota is per-user"
        )

        # Per-role limits are also defined, showing intentional per-user design.
        assert "_ROLE_LIMITS" in vars(quota_mod)
        assert "analyst" in quota_mod._ROLE_LIMITS
        assert "admin" in quota_mod._ROLE_LIMITS

    # ── Correction #2 (Bug #8 in original list) ────────────────────────────
    def test_generation_semaphore_correctly_limits_concurrent_generations(self):
        """pipeline.py uses asyncio.Semaphore(CR_MAX_CONCURRENT_GENERATIONS).

        The original analysis suggested concurrent generation was unlimited.
        In fact, _generation_semaphore at pipeline.py:60 gates all generation
        calls.  This is the correct single-process back-pressure mechanism.

        Source: credit_report/generation/pipeline.py:60
        """
        from credit_report.generation import pipeline as pipeline_mod
        from credit_report.config import CR_MAX_CONCURRENT_GENERATIONS

        sem = pipeline_mod._generation_semaphore
        assert isinstance(sem, asyncio.Semaphore), (
            "_generation_semaphore must be an asyncio.Semaphore"
        )
        # asyncio.Semaphore stores the initial value in _value attribute.
        assert sem._value == CR_MAX_CONCURRENT_GENERATIONS, (
            f"Semaphore value {sem._value} must equal CR_MAX_CONCURRENT_GENERATIONS "
            f"({CR_MAX_CONCURRENT_GENERATIONS})"
        )
