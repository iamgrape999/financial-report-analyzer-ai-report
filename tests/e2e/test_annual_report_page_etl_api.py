from __future__ import annotations

import io
import json
from pathlib import Path

import sys
ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from typing import Any

import httpx
import pytest

from credit_report.generation.document_pipeline import validate_section_proposal_gates
from credit_report.generation.evidence import _extract_text_from_xlsx

from .utils_e2e import (
    assert_no_server_error,
    count_section7_core_metrics,
    is_identity_only_payload,
    require_real_file,
    save_json_artifact,
)

pytestmark = pytest.mark.e2e


def _upload_annual_report(api_client: httpx.Client, auth_headers: dict[str, str], report_id: str, pdf_path: Path, selected_type: str = "financial_statement") -> dict[str, Any]:
    require_real_file(pdf_path, "TSMC_ANNUAL_REPORT_PATH")
    with pdf_path.open("rb") as fh:
        response = api_client.post(
            f"/reports/{report_id}/documents",
            headers=auth_headers,
            data={"document_type": selected_type},
            files={"file": (pdf_path.name, fh, "application/pdf")},
        )
    assert_no_server_error(response)
    save_json_artifact(f"upload_{selected_type}_response.json", {"status_code": response.status_code, "body": response.json() if response.headers.get("content-type", "").startswith("application/json") else response.text})
    assert response.status_code == 201, f"Upload failed: {response.status_code} {response.text[:2000]}"
    return response.json()


def _get_proposals(api_client, auth_headers, report_id: str, doc_id: str | None = None) -> list[dict[str, Any]]:
    suffix = f"?doc_id={doc_id}" if doc_id else ""
    response = api_client.post(f"/reports/{report_id}/smart-import/proposals{suffix}", headers=auth_headers)
    assert_no_server_error(response)
    assert response.status_code == 200, f"Proposal listing failed: {response.status_code} {response.text[:2000]}"
    return response.json()


def _first_doc_id(body: dict[str, Any]) -> str:
    doc_id = body.get("id") or body.get("doc_id")
    assert doc_id, f"Upload response did not include document id: {body}"
    return doc_id


def test_a_real_tsmc_upload_is_coerced_to_annual_report(api_client, auth_headers, e2e_report, tsmc_annual_report_path):
    report_id = e2e_report["id"]
    upload_body = _upload_annual_report(api_client, auth_headers, report_id, tsmc_annual_report_path, selected_type="financial_statement")
    doc_id = _first_doc_id(upload_body)
    docs_response = api_client.get(f"/reports/{report_id}/documents", headers=auth_headers)
    assert_no_server_error(docs_response)
    assert docs_response.status_code == 200, docs_response.text[:2000]
    docs = docs_response.json()
    doc = next((d for d in docs if d["id"] == doc_id), None)
    save_json_artifact("annual_report_detection_response.json", {"selected_document_type": "financial_statement", "upload": upload_body, "listed_document": doc})
    assert doc, f"Uploaded document {doc_id} not returned by GET documents: {docs}"
    inferred = upload_body.get("document_type") or doc.get("document_type")
    assert inferred == "annual_report" and doc.get("document_type") == "annual_report", (
        f"TSMC annual report was not coerced to annual_report. upload={upload_body}, listed={doc}"
    )
    assert doc.get("etl_status") == "uploaded", "Upload may be pending extraction, but document_type must already be annual_report."


def test_b_legacy_etl_endpoint_auto_scans_or_explicitly_blocks_one_shot(api_client, auth_headers, e2e_report, tsmc_annual_report_path):
    report_id = e2e_report["id"]
    upload_body = _upload_annual_report(api_client, auth_headers, report_id, tsmc_annual_report_path, selected_type="other")
    doc_id = _first_doc_id(upload_body)
    response = api_client.post(f"/reports/{report_id}/documents/{doc_id}/etl", headers=auth_headers)
    assert_no_server_error(response)
    body = response.json() if response.headers.get("content-type", "").startswith("application/json") else {"text": response.text}
    save_json_artifact("legacy_etl_auto_scan_or_block_response.json", {"status_code": response.status_code, "body": body})

    if response.status_code in {409, 422}:
        detail = json.dumps(body.get("detail", body), ensure_ascii=False).lower()
        assert any(term in detail for term in ("page", "scan", "coverage", "annual-report", "annual report")), f"Blocking error is not actionable/page-first-related: {body}"
        return

    assert response.status_code == 200, f"Unexpected ETL response: {response.status_code} {response.text[:2000]}"
    assert body.get("document_type") == "annual_report", f"ETL did not preserve annual_report type: {body}"
    proposals = _get_proposals(api_client, auth_headers, report_id, doc_id)
    save_json_artifact("proposals_after_legacy_etl_endpoint.json", proposals)
    assert body.get("sections_extracted") or proposals, "Legacy endpoint returned success but no extracted sections/proposals."
    for proposal in proposals:
        if proposal.get("section_no") == 7 and proposal.get("status") == "ready_for_review":
            metric_count, found, missing = count_section7_core_metrics(proposal.get("proposed_json") or {})
            save_json_artifact("section7_ready_proposal_metric_check.json", {"metric_count": metric_count, "found": found, "missing": missing, "proposal": proposal})
            assert metric_count >= 5 and not is_identity_only_payload(proposal.get("proposed_json") or {}), "Section 7 identity-only/low-metric proposal was marked ready_for_review."


def test_c_scan_pages_coverage_gate(api_client, auth_headers, e2e_report, tsmc_annual_report_path):
    report_id = e2e_report["id"]
    doc = _upload_annual_report(api_client, auth_headers, report_id, tsmc_annual_report_path)
    doc_id = _first_doc_id(doc)
    response = api_client.post(f"/reports/{report_id}/documents/{doc_id}/scan-pages", headers=auth_headers)
    assert_no_server_error(response)
    assert response.status_code == 200, f"scan-pages failed: {response.status_code} {response.text[:2000]}"
    payload = response.json()
    save_json_artifact("scan_pages_response.json", payload)
    for key in ("total_pages", "processed_pages", "coverage_pct", "native_text_pages", "financial_pages_detected"):
        assert key in payload, f"scan-pages response missing {key}: {payload}"
    assert payload["total_pages"] > 0, f"scan-pages total_pages == 0: {payload}"
    assert payload["processed_pages"] == payload["total_pages"], f"scan-pages did not process every page: {payload}"
    assert payload["financial_pages_detected"] > 0, f"No financial pages detected in real TSMC annual report: {payload}"


def test_d_plan_etl_produces_section_plan(api_client, auth_headers, e2e_report, tsmc_annual_report_path):
    report_id = e2e_report["id"]
    doc = _upload_annual_report(api_client, auth_headers, report_id, tsmc_annual_report_path)
    doc_id = _first_doc_id(doc)
    scan = api_client.post(f"/reports/{report_id}/documents/{doc_id}/scan-pages", headers=auth_headers)
    assert_no_server_error(scan)
    assert scan.status_code == 200, scan.text[:2000]
    response = api_client.post(f"/reports/{report_id}/documents/{doc_id}/plan-etl", headers=auth_headers)
    assert_no_server_error(response)
    assert response.status_code == 200, f"plan-etl failed: {response.status_code} {response.text[:2000]}"
    payload = response.json()
    save_json_artifact("plan_etl_response.json", payload)
    target = payload.get("target_sections")
    assert isinstance(target, dict) and target, f"plan-etl did not return target_sections: {payload}"
    missing = [str(n) for n in range(1, 11) if str(n) not in target]
    assert not missing, f"plan-etl missing sections 1-10: {missing}; target_sections={target}"
    section7 = target.get("7") or target.get(7)
    status = json.dumps(section7, ensure_ascii=False).lower()
    assert section7 and ("supported" in status or "partial" in status), f"Section 7 must be supported/partial_supported for annual report: {section7}"


def test_e_identity_only_section7_proposal_gate_rejects_false_success():
    payload = {"company_name_zh": "台灣積體電路製造股份有限公司", "responsible_person": "魏哲家", "spokesperson": "..."}
    result = validate_section_proposal_gates(7, payload, {"passed": True, "missing": [], "coverage_score": 1.0})
    save_json_artifact("identity_only_gate_result.json", {"payload": payload, "gate": result})
    assert result["passed"] is False
    assert result["coverage_score"] < 0.5


def test_f_empty_section7_proposal_gate_rejects_ready_status():
    for payload in ({}, None):
        result = validate_section_proposal_gates(7, payload, {"passed": True, "missing": [], "coverage_score": 1.0})
        assert result["passed"] is False, f"Empty/null payload must not pass: payload={payload} result={result}"
        assert "no_extracted_data" in result.get("missing", [])


def test_g_full_section7_financial_payload_is_ready_for_review():
    payload = {
        "revenue": 2639000000000, "gross_profit": 1520000000000, "operating_income": 1280000000000,
        "net_income": 1100000000000, "eps": 42.5, "total_assets": 6500000000000,
        "total_liabilities": 2100000000000, "total_equity": 4400000000000,
        "operating_cash_flow": 1500000000000, "cash": 1800000000000,
    }
    result = validate_section_proposal_gates(7, payload, {"passed": True, "missing": [], "coverage_score": 1.0})
    save_json_artifact("full_financial_payload_gate_result.json", {"payload": payload, "gate": result})
    assert result["passed"] is True
    assert result["extracted_metric_score"] >= 0.5


@pytest.mark.asyncio
async def test_h_low_coverage_proposal_cannot_commit(api_client, auth_headers, e2e_report):
    from credit_report.database import AsyncSessionLocal
    from credit_report.generation.models import SectionImportProposal

    report_id = e2e_report["id"]
    async with AsyncSessionLocal() as db:
        proposal = SectionImportProposal(
            report_id=report_id,
            section_no=7,
            proposed_json={"company_name_zh": "台灣積體電路製造股份有限公司"},
            evidence_map={"source_pages": [1], "coverage_gate": {"passed": False}},
            coverage_score=0.0,
            missing_required_fields=["revenue", "net_income"],
            status="low_coverage_failed",
        )
        db.add(proposal)
        await db.commit()
        proposal_id = proposal.id
    response = api_client.post(f"/reports/{report_id}/smart-import/{proposal_id}/commit", headers=auth_headers)
    save_json_artifact("low_coverage_commit_response.json", {"status_code": response.status_code, "body": response.json() if response.headers.get("content-type", "").startswith("application/json") else response.text})
    assert response.status_code == 409, f"Low-coverage proposal must not commit: {response.status_code} {response.text[:2000]}"
    input_response = api_client.get(f"/reports/{report_id}/inputs/7", headers=auth_headers)
    assert input_response.status_code in {404, 200}
    if input_response.status_code == 200:
        metric_count, _, _ = count_section7_core_metrics(input_response.json().get("input_json") or {})
        assert metric_count == 0, "Low-coverage rejected proposal unexpectedly wrote financial SectionInput."


def test_i_ready_section7_proposal_commit_persists_to_section_input(api_client, auth_headers, e2e_report, tsmc_annual_report_path):
    report_id = e2e_report["id"]
    doc = _upload_annual_report(api_client, auth_headers, report_id, tsmc_annual_report_path)
    doc_id = _first_doc_id(doc)
    scan = api_client.post(f"/reports/{report_id}/documents/{doc_id}/scan-pages", headers=auth_headers)
    assert_no_server_error(scan)
    assert scan.status_code == 200, scan.text[:2000]
    section7 = api_client.post(f"/reports/{report_id}/documents/{doc_id}/etl-section/7", headers=auth_headers)
    assert_no_server_error(section7)
    save_json_artifact("section7_etl_response.json", {"status_code": section7.status_code, "body": section7.json() if section7.headers.get("content-type", "").startswith("application/json") else section7.text})
    proposals = _get_proposals(api_client, auth_headers, report_id, doc_id)
    save_json_artifact("proposals_after_etl.json", proposals)
    ready = [p for p in proposals if p.get("section_no") == 7 and p.get("status") == "ready_for_review"]
    assert ready, f"No real Section 7 ready_for_review proposal was produced; cannot fake commit success. proposals={proposals}"
    proposal = ready[0]
    metric_count, found, missing = count_section7_core_metrics(proposal.get("proposed_json") or {})
    save_json_artifact("section7_ready_before_commit_metrics.json", {"metric_count": metric_count, "found": found, "missing": missing, "proposal": proposal})
    assert metric_count >= 5, f"Ready Section 7 proposal lacks core financial metrics: found={found}, missing={missing}"
    assert (proposal.get("evidence_map") or {}).get("source_pages"), f"Ready proposal has no source page evidence: {proposal}"
    commit = api_client.post(f"/reports/{report_id}/smart-import/{proposal['id']}/commit", headers=auth_headers)
    assert_no_server_error(commit)
    save_json_artifact("committed_ready_proposal.json", {"status_code": commit.status_code, "body": commit.json() if commit.headers.get("content-type", "").startswith("application/json") else commit.text})
    assert commit.status_code == 200, f"Ready proposal commit failed: {commit.status_code} {commit.text[:2000]}"
    first = api_client.get(f"/reports/{report_id}/inputs/7", headers=auth_headers)
    assert first.status_code == 200, f"SectionInput missing after commit: {first.status_code} {first.text[:2000]}"
    save_json_artifact("section7_input_after_commit.json", first.json())
    with httpx.Client(base_url=str(api_client.base_url), timeout=api_client.timeout, follow_redirects=True) as reload_client:
        second = reload_client.get(f"/reports/{report_id}/inputs/7", headers=auth_headers)
    assert second.status_code == 200, f"SectionInput missing after reload/new client: {second.status_code} {second.text[:2000]}"
    payload = second.json().get("input_json") or {}
    save_json_artifact("section7_input_after_commit_reload.json", second.json())
    persisted_count, persisted_found, persisted_missing = count_section7_core_metrics(payload)
    assert persisted_count >= 5, f"Persisted SectionInput lost financial metrics: found={persisted_found}, missing={persisted_missing}, payload={payload}"
    post_commit_proposals = _get_proposals(api_client, auth_headers, report_id, doc_id)
    committed = next((p for p in post_commit_proposals if p.get("id") == proposal["id"]), None)
    assert committed and (committed.get("evidence_map") or {}).get("source_pages"), "Proposal evidence/source pages must remain queryable after commit."


def test_j_xlsx_merged_cells_and_cap_regression():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged Cap"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Revenue"
    ws["A2"] = "FY2025"
    ws["B2"] = 123
    for row in range(1, 61):
        for col in range(1, 61):
            if row == 1 and col in (1, 2):
                continue
            ws.cell(row=row, column=col).value = ws.cell(row=row, column=col).value or f"R{row}C{col}"
    ws["A55"] = "SHOULD_NOT_BE_EXTRACTED_ROW_55"
    ws.cell(row=2, column=55).value = "SHOULD_NOT_BE_EXTRACTED_COL_55"
    buf = io.BytesIO()
    wb.save(buf)
    text = _extract_text_from_xlsx(buf.getvalue())
    assert "Revenue" in text, "Merged-cell anchor value was not preserved."
    assert "FY2025" in text and "123" in text, "Expected data cells missing from xlsx extraction."
    assert "SHOULD_NOT_BE_EXTRACTED_ROW_55" not in text, "Rows beyond cap were extracted as if complete."
    assert "SHOULD_NOT_BE_EXTRACTED_COL_55" not in text, "Columns beyond cap were extracted as if complete."
    assert len(text) < 80_000, f"XLSX extraction grew unexpectedly large: {len(text)} chars"
